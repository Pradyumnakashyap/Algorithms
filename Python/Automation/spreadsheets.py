#Access the spreadsheet with the openpyxl module and perform some functions


import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import glob as g


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2,sheet.max_row + 1):    # provides the max number of rows in the sheet, iterate on those
        cell = sheet.cell(row, 3)              #third column has value required, so hard code with iterative values
        updated_price = cell.value * 0.9
        updated_price_cell = sheet.cell (row,4)
        updated_price_cell.value = updated_price

    values = Reference(sheet, min_row=2,                        #Reference object gets reference to specific cells
                            max_row=sheet.max_row, 
                            min_col=4, 
                            max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    wb.save(filename)

files = g.glob('/home/pk/Desktop/Python/Algorithms/Python/Automation/*.xlsx')

for file in files:
    process_workbook(file)
    
